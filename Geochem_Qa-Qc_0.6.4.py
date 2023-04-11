"""
##############################################################################
Copyright 2022 Commonwealth of Australia (Geoscience Australia)
 
Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at
 
http://www.apache.org/licenses/LICENSE-2.0
 
Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.
##############################################################################
"""
import sys
import os

import numpy as np
import pandas as pd
import statistics as st
import seaborn as sns
import matplotlib.pyplot as plt
import matplotlib.backends.backend_pdf
import xlsxwriter
from openpyxl import load_workbook

'''
This section covers the user defined variables for each run. It is required
to set the file_name, save_location, Id_column, and duplicate/replicate names.

'''

FILE_NAME = r""
Save_Location = r""
# The name of the column that contains the sample numbers
Id_Coloumn = ''
# minimum number of times a sample repeats before been included as a standard
STANDARD_CUTOFF = 3
# name of the lab duplicates
DUPLICATE_NAME = ''
# name of the analytical duplicates
REPLICATE_NAME = ''
# Enable Debugging
DEBUG = False
# Does the data contain multpile batches, False = No, True = yes
BATCHED = False
BATCH = ''
# if the data is batched, specify the row numbers for the start of each batch
BATCHES = []

def parse(geochem_data):
    '''
    Parses the header information in order to find geochemical elements and
    oxides. The function will return a cut down version of the header and the
    data to just include the geochemical data.

    Parameters
    ----------
    geochem_data : dataframe
        A pandas dataframe containing geochemistry data with headers to be
        parsed.

    Returns
    -------
    element_list : array
        An array containing the elements and oxides within the dataset.
    geochem_data : dataframe
        The input dataframe returned with the header information updated to
        searchable .

    '''
    common_headers = ("Lat", 'Long', 'Latitude', 'Longitude', 'Lab','Sample',
              'Time','Date', 'Group','Elev', 'Type', 'Site', 'Comment',
              'Depth','Size', 'LAT', 'LONG', 'Lab No', 'STATE', 'majors',
              'Recode','Name', 'East','North', 'LOI', 'SAMPLE', "GRAIN",
              "BATCH","Survey", "ID", "Standard", "Sample", "Colour", "batch",
              "sampleno", "SampleID", "Sampleno", "Jobno", "Pair", "Order",
              "Internal", "External", "METHOD", "SampleNo", 'Sample No',
              'Sample ID', 'External Lab No.', 'Internal Lab No.', 'Batch',
              'METHOD MILL', 'GA Sample No.', 'ENO', 'SITEID', 'LATITUDE',
              'LONGITUDE', 'BV_ID', 'Pair', 'Batch', 'Order', 'Chem',
              'Sampleid', 'LabNo', 'Acq')
    elements = ('SiO2', 'TiO2','Al2O3', 'Fe2O3', 'FeO','MnO', 'MgO', 'CaO',
                'Na2O','K2O', 'P2O5', 'SO3', "H", "He", "Li", "Be", "B", "C",
                "N","O", "F", "Ne", 'Na', 'Mg', 'Al', 'Si','P', 'S', 'Cl',
                'Ar','K', 'Ca', 'Sc', 'Ti', 'V', 'Cr', 'Mn', 'Fe', 'Co', 'Ni',
                'Cu', 'Zn', 'Ga', 'Ge', 'As', 'Se', 'Br', 'Kr', 'Rb', 'Sr',
                'Y', 'Zr','Nb', 'Mo', 'Tc', 'Ru', 'Rh', 'Pd', 'Ag', 'Cd',
                'In', 'Sn', 'Sb', 'Te','Xe', 'Cs', 'Ba', 'La', 'Ce', 'Pr',
                'Nd','Pm','Sm', 'Eu', 'Gd', 'Tb', 'Dy', 'Ho', 'Er','Tm', 'Yb',
                'Lu', 'Hf', 'Ta', 'W', 'Re', 'Os', 'Ir', 'Pt', 'Au', 'Hg',
                'Tl','Pb', 'Bi', 'Po', 'At', 'Rn', 'Fr', 'Ra', 'Ac', 'Th',
                'Pa', 'U', 'Np', 'Pu','Am', 'Cm', 'Bk', 'Cf', 'Es', 'Fm',
                'Md', 'No', 'Lr', 'Rf', 'Db', 'Sg', 'Bh', 'Hs', 'Mt', 'LOI',
                'I', 'ORGC')
    header = list(geochem_data)
    index = []
    element_list = []
    for i, value in enumerate(header):
        if any(x in value for x in elements):
            if value not in common_headers:
                index.append(i)
            elif value in common_headers[2]:
                header[i] = 'Latitude'
        elif value in common_headers[3]:
            header[i] = 'Longitude'
    for i in index:
        temp_store = []
        for count, value in enumerate(elements):
            if value in header[i]:
                if len(value) > len(temp_store):
                    temp_store = value
        header[i] = temp_store
        element_list.append(temp_store)
    geochem_data.columns = header
    return element_list, geochem_data


def LLD(geochem_data, element_list, imputation = False):
    '''
    This function is designed to find values that are below the detection
    limit of the analytical method. The function works by searching for the <
    symbol in the

    Parameters
    ----------
    geochem_data : dataframe
        Dataframe containg the geochemical data.
    element_list : array
        An array containing the elements found in the dataframe.
    imputation : boolean, optional
        currently not implemented. The default is False.

    Returns
    -------
    geochem_data : Dataframe
        The input dataframe with values at the limit of detction set to
        0.5*detection limit.
    detection_limits : array
        A numpy array contianing the detected  limit of decection for each
        element.

    '''
    detection_limits = np.zeros([len(element_list)])
    for i in range (len(element_list)):
        less_than = False
        # find the values that have a < and turn that into a value
        if (geochem_data[element_list[i]].apply(str).\
                str.contains('<', na=False,regex=True).any()) == True:
            index = geochem_data[element_list[i]].\
                str.contains('<', na=False,regex=True)
            less_than = True
        else:
            index = geochem_data[element_list[i]].apply(str).\
                str.contains('-', na=False,regex=True)
            #except:
               # pass
        detection_index = geochem_data[element_list[i]].loc[index].unique()
        detection_index = list(detection_index)
        if len(detection_index) > 0:
            detection_index = detection_index[0]
            if less_than == True:
                new_detection_index = detection_index.replace('<','')
                new_detection_index = float(new_detection_index)
            else:
                new_detection_index = detection_index
            if new_detection_index <0:
                dl_value = new_detection_index*-1
            else:
                dl_value = new_detection_index
            #create a half detection limit value
            detection_limits[i] = dl_value
            half_detection = float(dl_value)/2
            # replace the less than values with half detection limits
            geochem_data[element_list[i]] = \
                geochem_data[element_list[i]].replace(detection_index,
                                                      half_detection)
    if imputation == True:
        print('Still to be implemented')
    return geochem_data, detection_limits

def repeats(geochem_data):
    '''
    Function used to find the location of repeats within a dataframe using a
    key. The pair is presumed to be the location -1.

    Parameters
    ----------
    geochem_data : dataframe
        The dataframe containing the full geochemical dataset.

    Returns
    -------
    rep_location : list
        The location of the repeats as identified by the key.
    rep_pair : list
        The location of the corresponding pair for each of the identified
        repeats.

    '''
    # Find the location of repeats and their pair
    try:
        repeat = geochem_data[geochem_data[Id_Coloumn].str.contains(" R",
                                                                     na=False)]
    except KeyError:
        print ("Unable to find a column with the name: ", Id_Coloumn)
        sys.exit()
    rep_location = list(repeat.index)
    rep_pair = [x-1 for x in rep_location]
    return rep_location, rep_pair

def standard_stats(geochem_data, element_list, detection_limits = False):
    '''
    This Function is used to calculate the summary statistics for the analysed
    standards. The main statistics calculated are: mean, standard deviation,
    RSD (relative standard deviation). For analyses with more than one
    standard, the weighted average of the rsds is calculated using the
    mean concentration as a weighting system.

    Parameters
    ----------
    geochem_data : dataframe
        Dataframe containing  the geochemical data.
    element_list : list
        List containing the elements present within the dataframe.

    Returns
    -------
    None.

    '''
    stds_count = geochem_data[Id_Coloumn].value_counts()
    print(stds_count)
    Stds = stds_count[stds_count > STANDARD_CUTOFF]
    Stds_List = Stds.index.tolist()
    # Gernerate the headers for the Standard statistic
    for i in range (0, len(Stds_List)):
        if i == 0:
            stats_header = Stds_List[i] + '_Average'
        else:
            stats_header = np.append(stats_header,str(Stds_List[i]) +
                                     '_Average')
        stats_header = np.append(stats_header, [str(Stds_List[i]) + '_StDev',
                                                str(Stds_List[i]) +
                                                '_rsd(%)'])
    # Create a pandas dataframe to store the statitics
    # loop to add in a weighted average for more than one standard
    if len(Stds_List) >1:
        stats_header = np.append(stats_header, 'Weighted_Average')
        temp_stats = np.zeros([len(element_list),(len(Stds_List)*3)+1])
        average_locations = np.arange(0,len(Stds_List)*3, step = 3)
        rstdev_locations = [x+2 for x in average_locations]
    else:
        temp_stats = np.zeros([len(element_list),len(Stds_List)*3])
    std_stats = pd.DataFrame(data=temp_stats,
                             index=element_list,
                             columns= stats_header)
    pdf = matplotlib.backends.backend_pdf.PdfPages(Save_Location +
                                                   "\\Standards.pdf")
    try:
        #create a folder for the standrd images
        os.mkdir(Save_Location + "\\Standards")
    except FileExistsError:
        print ("Folder already exists")
    colour = ('tab:blue', 'tab:orange', 'tab:green', 'tab:red', 'tab:purple',
              'tab:cyan','tab:gray','tab:pink','tab:olive', 'dodgerblue')
    print (Stds_List)
    for j in range(0, len(element_list)):
        xlimit = 0
        y_upper = 0
        y_lower = 0
        for i in range(0, len(Stds_List)):
            element = geochem_data.loc[geochem_data[Id_Coloumn] ==
                                       Stds_List[i],element_list[j]]
            if min(element) < y_lower:
                y_lower = min(element)
            if max(element) > y_upper:
                y_upper = max(element)
            #option to turn on/off
            #element.index = np.arange(start=1, stop=(len(element)+1), step=1)
            element.columns = Stds_List[i]
            std_stats.loc[element_list[j],
                          stats_header[i+(2*i)]] = element.mean()
            '''
            need to look into if the mean is used and what impact it is having
            on the data

            '''
            std_stats.loc[element_list[j],
                          stats_header[(i+(2*i))+1]] = element.std()
            std_stats.loc[element_list[j],
                          stats_header[(i+(2*i))+2]] = (element.std()/
                                                        element.median())*100
            splot = sns.scatterplot(data=element, color = colour[i],
                                    label =  Stds_List[i])
            if max(element.index) > xlimit:
                xlimit = (max(element.index))
            plt.plot((0,99999),(element.median(),element.median()),
                     linewidth=0.5, color = colour[i])
            plt.plot((0,99999),((element.median())*1.1,(element.median())*1.1),
                     linewidth=0.5,linestyle='dashed', color = colour[i])
            plt.plot((0,99999),((element.median())*0.9,(element.median())*0.9),
                     linewidth=0.5, linestyle='dashed', color = colour[i])
        savename = ('{}\\Standards\\{}_Standards.png').format(Save_Location,
                                                              element_list[j])
        lgd = splot.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
        plt.xlim(0,xlimit)
        if detection_limits[j] > 0:
            plt.plot((0,99999),(detection_limits[j],detection_limits[j]),
                     linestyle='dashed', linewidth=1, color = 'k')
        if BATCHED:
            for i in range(len(BATCHES)):
                plt.plot((BATCHES[i] + 0.5, BATCHES[i] + 0.5), (-99999,99999),
                         color = 'k', linestyle='dashed', linewidth = 0.5)
            plt.ylim(y_lower*1.3, y_upper*1.3)
        plt.title(element_list[j])
        plt.savefig(savename, bbox_extra_artists=(lgd,),
                    bbox_inches='tight', format = 'png',
                    orientation='portrait',dpi = 900)
        pdf.savefig()
        plt.close()
        plt.clf()
    pdf.close()
    if len(Stds_List) >1:
        for i in range(0, len(element_list)):
            equation_temp = 0
            sum_stds = std_stats.loc[element_list[i],
                                     stats_header[average_locations]].sum()
            for j in range (0,len(average_locations)):
                avg = std_stats.loc[element_list[i],
                                    stats_header[average_locations[j]]]
                rsd = std_stats.loc[element_list[i],
                                    stats_header[rstdev_locations[j]]]
                equation_temp = equation_temp + (rsd*avg)
            std_stats.loc[element_list[i],
                          'Weighted_Average'] = equation_temp/sum_stds
    standards_save = ('{}\\Standard_Stats.xlsx').format(Save_Location)
    try:
        std_stats.to_excel(standards_save)
    except xlsxwriter.exceptions.FileCreateError:
        print ('''Unable to save, the file must already be open. A new verison
               will try to be saved with _new. If this fails no file will be
               saved.''')
        standards_save = ('{}\\Standard_Stats_new.xlsx').format(Save_Location)
        std_stats.to_excel(standards_save)
    del pdf


def lab_assessment(geochem_data, element_list, detection_limits):
    '''
    We need to calculate Min, MAX, median and 90th percentile RSD (plus number
    ) for pairs for the following intervals:
    1-5* DL - group 1
    5-10* DL - group 2
    10-50* DL - group 3
    50-100* DL - group 4
    100-500* Dl - group 5
    500-1000* DL- group 6
    >1000* DL - group 7


    Parameters
    ----------
    geochem_data : TYPE
        DESCRIPTION.
    elements_list : TYPE
        DESCRIPTION.

    Returns
    -------
    None.

    '''
    try:
        detection_limits  = pd.read_excel(detection_limits, header = 0)
    except FileNotFoundError:
        print ('Could not find a file with that name')
        sys.exit()
    #count if to determine groups
    repeat_loation, repeat_Pair = repeats(geochem_data)
    dl = detection_limits
    detection_limits = detection_limits[element_list]
    dl = list(detection_limits)
    pdf = matplotlib.backends.backend_pdf.PdfPages(Save_Location +
                                               "\\Laboratory_Assessment.pdf")
    try:
        os.mkdir(Save_Location + "\\Laboratory_Assessment")
    except FileExistsError:
        print ("Folder already exists")
    for i in range(0, len(element_list)):
        dl = detection_limits[element_list[i]]
        dl = list(dl)
        x = geochem_data.iloc[repeat_loation,list(
            geochem_data.columns).index(element_list[i])]
        y = geochem_data.iloc[repeat_Pair,list(
            geochem_data.columns).index(element_list[i])]
        means = np.zeros([len(x)])
        rpt = list(x)
        rpt_pair = list(y)
        for j in range(len(x)):
            means[j] = np.mean((rpt[j], rpt_pair[j]))
        #find the number of samples in each group
        group_count = np.zeros([7])
        group_ranges = [1,5,10,50,100,500,1000,9999999999]
        for j in range(7):
            group_count[j] = sum(1 for x in means if x> group_ranges[j]*dl[0]
                                 and x< group_ranges[j+1]*dl[0])
        temp_summary_stats = np.zeros([4,7])
        stats_header = ('1-5* DL', '5-10* DL', '10-50* DL', '50-100* DL',
                        '100-500* Dl', '500-1000* DL', '>1000* DL')
        stats_index = ('Minimum','Maximum','Median','90th percentile')
        summary_stats = pd.DataFrame(data=temp_summary_stats,
                                     index=stats_index,columns= stats_header)
        for j in range(7):
            if group_count[j] > 1:
                repeat  = np.array(rpt)
                repeat_pair = np.array(rpt_pair)
                repeat_index = [x for x, y in enumerate(means) if
                                y> group_ranges[j]*dl[0]and
                                y< group_ranges[j+1]*dl[0]]
                if DEBUG:
                    print (group_count[j], 'number in group')
                    print (len(repeat_index), 'number found')
                    print (group_ranges[j]*dl[0], group_ranges[j+1]*dl[0])
                    print (dl, 'detection limit')
                    print (means[repeat_index], 'found values')
                    print (repeat[repeat_index], 'repeats')
                    print (repeat_pair[repeat_index], 'repeat pairs')
                    print ('---------------------')
                #summary_stats[stats_header[j]] = duplicate_statistics(
                    #repeat[repeat_index], repeat_pair[repeat_index])
                #minimum, maximum, median, percentile =  duplicate_statistics(
                    #repeat[repeat_index], repeat_pair[repeat_index])
            else:
                summary_stats[stats_header[j]] = ['NaN','NaN','NaN','NaN']
        stats_save = ('{}\\Laboratory_Assessment_Stats.xlsx').format(
            Save_Location)
        if i == 0:
            with pd.ExcelWriter(stats_save, engine='xlsxwriter') as writer:
                summary_stats.to_excel(writer, element_list[i])
                writer.save()
        book = load_workbook(stats_save)
        with pd.ExcelWriter(stats_save, engine='openpyxl') as writer:
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            summary_stats.to_excel(writer, element_list[i])
            writer.save()
        plt.scatter(x, y, s= 3)
        plt.title(element_list[i])
        colour = ['r','cornflowerblue','forestgreen','orange',
                  'teal','darkmagenta','k']
        for count in range(7):
            detection_line = group_ranges[count]*dl[0]
            plt.plot([detection_line,detection_line],[0,999999],
                     linestyle='dashed', color = colour[count])
            plt.plot([0,999999],[detection_line,detection_line],
                     linestyle='dashed', color = colour[count])
        x_min = min(x)-(min(x)*0.2)
        y_min = min(y)-(min(y)*0.2)
        if x_min <0:
            x_min = 0
        if y_min <0:
            y_min = 0
        plt.xlim(x_min,max(x)+(max(x)*0.2))
        plt.ylim(y_min,max(y)+(max(y)*0.2))
        plt.yscale('log')
        plt.xscale('log')
        savename = ('{}\\Laboratory_Assessment\\{}_Laboratory_Assessment.png'
                    ).format(Save_Location,element_list[i])
        plt.savefig(savename, format = 'png', orientation='portrait',
                    dpi = 900)
        pdf.savefig()
        plt.close()
        plt.clf()
    writer.close()
    pdf.close()
                #for count in range(group_count[j]):

                #use repeat and repeat pair and duplicate_statistics function

def Duplicates(geochem_data, key_word, SheetName, element_list):
    '''
   Function use to locate each of the duplicate pairs. Once found statistics
   are performed on each pair and a linear regression plot produced.

    Parameters
    ----------
    geochem_data : Dataframe
        DESCRIPTION.
    key_word : String
        DESCRIPTION.
    SheetName : String
        DESCRIPTION.

    Returns
    -------
    None.

    '''
    repeat = geochem_data[geochem_data[Id_Coloumn].str.contains(key_word,na=False)]
    rep_location = list(repeat.index)
    repeat_list = list(repeat[Id_Coloumn])
    repeat_list = [item.replace(key_word, "") for item in repeat_list]
    duplicates = pd.DataFrame(np.zeros((len(rep_location)*2, len(list(geochem_data)))),columns = list(geochem_data))
    #duplicates.loc[0] = geochem_data.loc[0]
    for i in range(0,len(rep_location)):
        try:
            duplicates.loc[i*2] = geochem_data.loc[rep_location[i]]
           # try:
            #repeat_location = geochem_data[geochem_data[Id_Coloumn].str.contains(repeat_list[i],na=False)]
            try:
                repeat_location = geochem_data[(geochem_data[Id_Coloumn]) == int(repeat_list[i])]
                if len(repeat_location) <1:
                    repeat_location = geochem_data[(geochem_data[Id_Coloumn]) == str(repeat_list[i])]
            except:
                repeat_location = geochem_data[(geochem_data[Id_Coloumn]) == str(repeat_list[i])]
            repeat_location = (list(repeat_location.index))
            duplicates.loc[(i*2)+1] = geochem_data.loc[repeat_location[0]]
        except IndexError:
            pass
        #     pass
        #     print ("Duplicates funciton did not run correctly")
    duplicate_save = ('{}\\{}_Duplicates.xlsx').format(Save_Location, BATCH)
    duplicate_plotting(duplicates, element_list, SheetName)
    file_exists = os.path.isfile(duplicate_save)
    if file_exists == False:
        duplicates.to_excel(duplicate_save, index = False,sheet_name = SheetName)
        book = load_workbook(duplicate_save)
        writer = pd.ExcelWriter(duplicate_save, engine = 'openpyxl')
        writer.book = book
    else:
        book = load_workbook(duplicate_save)
        writer = pd.ExcelWriter(duplicate_save, engine = 'openpyxl')
        writer.book = book
        duplicates.to_excel(writer, index = False,sheet_name = SheetName)
    stats = Duplicate_Statistics(duplicates)
    stats.to_excel(writer, index = False,sheet_name = SheetName+"_Statistics")
    writer.save()
    writer.close()

def Duplicate_Statistics(duplicates):
    '''
    Calculates the statistics for duplicate pairs

    Parameters
    ----------
    duplicates : TYPE
        A pandas dataframe containing the duplicate pairs, generated by the
        Duplicates function.

    Returns
    -------
    duplicate_stats : array
        Array containing the statistics for each of the duplicate pairs.

    '''
    element_list, duplicates = parse(duplicates)
    number_duplicates = int(len(duplicates)/2)
    stats = np.zeros([number_duplicates, 4])
    duplicate_stats = pd.DataFrame(np.zeros([number_duplicates*5,len(element_list)+1]))
    header_info = np.append("",element_list)
    duplicate_stats.columns = header_info
    for j in range(len(element_list)):
        for i in range(number_duplicates):
            repeat = duplicates[element_list[j]].iloc[i*2]
            pair = duplicates[element_list[j]].iloc[(i*2)+1]
            stats[i,0] = st.mean([repeat, pair])
            stats[i,1] = st.stdev([repeat, pair])
            dup_id = duplicates.columns.get_loc(Id_Coloumn)
            duplicate_stats.iloc[(i*5),0] = duplicates.iloc[i*2,dup_id]
            duplicate_stats.iloc[(i*5)+1,0] =  "Mean"
            duplicate_stats.iloc[(i*5)+2,0] = "Standard Deviation"
            duplicate_stats.iloc[(i*5)+3,0] = "RSD"
            duplicate_stats.iloc[(i*5)+4,0] = "RPD"
            duplicate_stats.iloc[(i*5),j+1] = ""
            duplicate_stats.iloc[(i*5)+1,j+1] =  st.mean([repeat, pair])
            duplicate_stats.iloc[(i*5)+2,j+1] = st.stdev([repeat, pair])
            duplicate_stats.iloc[(i*5)+3,j+1] = stats[i,1]/stats[i,0]
            duplicate_stats.iloc[(i*5)+4,j+1] = abs(((repeat-pair)/((repeat+pair)/2))*100) #Relative Percent Difference
    return duplicate_stats

def duplicate_plotting(duplicates, element_list, name):
    '''
    Plots each of the duplicate pairs on an X-Y plot

    Parameters
    ----------
    duplicates : TYPE
        DESCRIPTION.
    element_list : TYPE
        DESCRIPTION.
    name : TYPE
        DESCRIPTION.

    Returns
    -------
    None.

    '''
    try:
        os.mkdir(Save_Location + "\\"+name)
    except FileExistsError:
        print ('Directory already exists')
    lengh = int(len(duplicates)/2)
    x_index = np.zeros([lengh])
    y_index = np.zeros([lengh])
    for i in range(lengh):
        x_index[i] = int(i*2)
        y_index[i] = int((i*2)+1)
    x = duplicates.loc[x_index]
    print (x)
    y = duplicates.loc[y_index]
    for element in element_list:
        print (element)
        x_data = x[element].to_numpy()
        y_data = y[element].to_numpy()
        n = len(x_data)
        delete_list = []
        for i in range(n):
            if type(x_data[i]) is str or type(y_data[i]) is str:
                delete_list.append(i)
        if len(delete_list) > 0:
            x_data = np.delete(x_data, delete_list)
            y_data = np.delete(y_data, delete_list)
        slope, intercept, r2 = linreg(x_data,y_data)
        fitx = np.linspace(-5,200000,100)
        fity = slope * fitx + intercept
        plt.plot(fitx, fity, color = 'lightcoral')
        plt.plot([0,9999999999],[0,9999999999],color = 'k',
                  linestyle='dashed')
        plt.xlim(min(x_data)-(min(x_data)*0.1),
                  max(x_data)+(max(x_data)*0.1))
        plt.ylim(min(y_data)-(min(y_data)*0.1),
                  max(y_data)+(max(y_data)*0.1))
        plt.scatter(x_data, y_data, s = 1)
        plt.title(element)
        plt.xlabel('Duplicate')
        plt.ylabel('Original')
        plot_title = ('N = ' + str(len(x_data)) + '\ny = ' +
                  str(round(slope, 2)) + 'x + ' + str(round(intercept, 2))+
                  '\nR$\mathregular{^2}$ = ' + str(round(r2,3)))
        plt.legend(title=plot_title)
        savename = r"{}\{}-LR.png".format(Save_Location + "\\"+name,element)
        plt.savefig(savename, format = 'png', dpi = 900)
        plt.close()
        plt.clf()

def linreg(x,y):
    """
    Calculates the slope, intercept, and r2 of two datasets using ordinary
    least squares linear regression.

    Parameters
    ----------
    x : array
        The data to be used as for the X-axis.
    y : array
        The data to be used as for the Y-axis.

    Returns
    -------
    m : float
        Slope of the linear regression.
    c : float
        Intercept of the linear regression.
    r2 : float
        R2 of the linear regression.

    """
    # confirm that the arrays contain valid data
    n = len(x)
    i = 0
    xy = np.zeros([n])
    x2 = np.zeros([n])
    for i in range(0,n):
        xy[i] = x[i]*y[i]
        x2[i] = x[i]**2
    m = (n*(sum(xy))-(sum(x)*sum(y)))/(n*(sum(x2))-(sum(x))**2)
    c = (sum(y)-m*sum(x))/n
    # R2 of the regression
    yscore = (1/len(y))*sum(y)
    sstot = np.zeros([len(y)])
    f = np.zeros([len(y)])
    ssres = np.zeros([len(y)])
    for i in range(0,len(y)):
        sstot[i] = (y[i] - yscore)**2
        f[i] = (m*x[i]) + c
        ssres[i] = (y[i] - f[i])**2
    sstot = sum(sstot)
    ssres = sum(ssres)
    r2 = 1-(ssres/sstot)
    return m, c, r2
#

def main():
    try:
        geochem_data  = pd.read_excel(FILE_NAME, header = 0)
    except FileNotFoundError:
        print ('Could not find a file with that name')
        sys.exit()
    #
    #header = list(geochem_data)
    element_list, geochem_data = parse(geochem_data)
    print (element_list)
    geochem_data, detection_limits = LLD(geochem_data,element_list)
    geochem_data_unstripped = geochem_data.copy(deep = True)
    geochem_data[Id_Coloumn] = geochem_data[Id_Coloumn].replace(' Rpt', '',
                                                               regex=True)
    # creates the plots and runs standard statistics
    print(detection_limits)
    standard_stats(geochem_data,element_list, detection_limits)
    if len(DUPLICATE_NAME) >0:
        Duplicates(geochem_data, DUPLICATE_NAME, "Lab_Duplicates",element_list)
    if len(REPLICATE_NAME) >0:
        Duplicates(geochem_data_unstripped, REPLICATE_NAME, "Analytical_Duplicates",element_list)
    #creates the plots for duplicate pairs
    #duplicates(geochem_data_unstripped, element_list)
    #lab_assessment(geochem_data_unstripped, element_list, detection_limits)

if __name__ == "__main__":
    main()

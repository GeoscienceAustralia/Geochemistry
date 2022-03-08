"""
Created on Wed Dec  2 14:37:12 2020

"""
import sys
import os

import numpy as np
import pandas as pd
import statistics as st
import seaborn as sns
import matplotlib.pyplot as plt
import matplotlib.backends.backend_pdf
import xlsxwriter
from openpyxl import load_workbook

FILE_NAME = r"C:\Users\u29043\Desktop\AS_Stats\Fire Assay\u328926_Final_AS_edits.xlsx"
Save_Location = r"C:\Users\u29043\Desktop\AS_Stats\Fire Assay"
Id_Coloumn = 'SampleID' # The name of the column that contains the sample numbers
DEBUG = 0
BATCHED = 0 # Does the data contain multpile batches, 0 = No 1 = Yes
BATCH = 'Batch 1'
BATCHES = [249,467] # if the data is batched, specify the row numbers for the start of each batch
STANDARD_CUTOFF = 2 #minimum number of times a sample repeats before been included as a standard

def parse(header):
    '''
    Parses the header informaiton in order to find geochemical elements and
    oxides. The function will return a cut down version of the header and the
    data to just include the geochemical data.

    Parameters
    ----------
    header : array of strings
        The array containing headers for each of the coloumns.

    Returns
    -------
    sndhead : array of strings
        A cut down header containing only chemical elements.
    fullhead : array of strings
        The full header, with Longitude and Latitude changed to a standardised
        name and the elements changed to standard notation.

    '''
    lengh = len(header)
    newhead = np.zeros([lengh], dtype = "U16")
    # list of common headers, add things here if you need them
    common = ("Lat", 'Long', 'Latitude', 'Longitude', 'Lab','Sample', 'Time',
              'Date', 'Group','Elev', 'Type', 'Site', 'Comment', 'Depth',
              'Size', 'LAT', 'LONG', 'Lab No', 'STATE', 'majors', 'Recode',
              'Name', 'East','North', 'LOI', 'SAMPLE', "GRAIN", "BATCH",
              "Survey", "ID", "Standard", "Sample", "Colour", "batch",
              "sampleno", "SampleID", "Sampleno", "Jobno", "Pair", "Order",
              "Internal", "External", "METHOD")
    elements = ('SiO2', 'TiO2','Al2O3', 'Fe2O3', 'FeO','MnO', 'MgO', 'CaO', 'Na2O',
                'K2O', 'P2O5', 'SO3', "H", "He", "Li", "Be", "B", "C", "N",
                "O", "F", "Ne", 'Na', 'Mg', 'Al', 'Si','P', 'S', 'Cl', 'Ar',
                'K', 'Ca', 'Sc', 'Ti', 'V', 'Cr', 'Mn', 'Fe', 'Co', 'Ni',
                'Cu', 'Zn', 'Ga', 'Ge', 'As', 'Se', 'Br', 'Kr', 'Rb', 'Sr',
                'Y', 'Zr','Nb', 'Mo', 'Tc', 'Ru', 'Rh', 'Pd', 'Ag', 'Cd',
                'In', 'Sn', 'Sb', 'Te','Xe', 'Cs', 'Ba', 'La', 'Ce', 'Pr',
                'Nd','Pm','Sm', 'Eu', 'Gd', 'Tb', 'Dy', 'Ho', 'Er','Tm', 'Yb',
                'Lu', 'Hf', 'Ta', 'W', 'Re', 'Os', 'Ir', 'Pt', 'Au', 'Hg',
                'Tl','Pb', 'Bi', 'Po', 'At', 'Rn', 'Fr', 'Ra', 'Ac', 'Th',
                'Pa', 'U', 'Np', 'Pu','Am', 'Cm', 'Bk', 'Cf', 'Es', 'Fm',
                'Md', 'No', 'Lr', 'Rf', 'Db', 'Sg', 'Bh', 'Hs', 'Mt', 'LOI',
                'I')
    for i in range(0, lengh):
        check = 0
        for k in range(0, len(common)):
            if common[k] in header[i]:
                newhead[i] = 'NaN'
                check = 1
            elif k == (len(common)-1) and check == 0:
                newhead[i] = header[i]
    nan_count = sum(1 for item in newhead if item==('NaN'))
    sndhead = np.zeros([len(newhead)-nan_count], dtype = "U16")
    fullhead = np.zeros(len(header), dtype = 'U64')
    for i in range(0,len(fullhead)):
        fullhead[i] = header[i]
    for i in range(0, len(fullhead)):
        if common[0] in fullhead[i] or common[2] in fullhead[i]:
            fullhead[i] = "Latitude"
        elif common[1] in fullhead[i] or common[3] in fullhead[i]:
            fullhead[i] = "Longitude"
    tstore = 'a'
    for i in range(0, len(newhead)):
        check = 0
        if newhead[i] != 'NaN':
            for k in range(0,len(elements)):
                if elements[k] in fullhead[i]:
                    if check == 0:
                        tstore = elements[k]
                        check = 1
                    elif check == 1 and len(tstore) ==1 and elements[k] !='I':
                        tstore = elements[k]
            fullhead[i] = tstore
    nan_count = sum(1 for item in newhead if item==('NaN'))
    delrows = np.zeros([nan_count])
    count = 0
    k = 0
    for i in range(0, len(newhead)):
        if newhead[i] =='NaN':
            delrows[count] = i
            count = count + 1
        else:
            sndhead[k] = newhead[i]
            k = k + 1
    newhead = np.zeros([len(sndhead)], dtype = "U16")
    tstore = "a"
    for i in range(0, len(sndhead)):
        check = 0
        for k in range(0, len(elements)):
            if elements[k] in sndhead[i]:
                #print elements[K], sndhead[I]
                if check == 1:
                    if len(tstore) == 1 and len(elements[k]) >1:
                        tstore = elements[k]
                elif check == 0:
                    tstore = elements[k]
                check = 1
                newhead[i] = tstore
            if check == 0 and k == 107:
                print (sndhead[i])
                newhead[i] = 'NaN'
                print ("element not found")
    nan_count = sum(1 for item in newhead if item==('NaN'))
    if nan_count == 0:
        sndhead = newhead
    else:
        sndhead = np.zeros([len(newhead)-nan_count], dtype = "U16")
        k = 0
        for i in range(0, len(newhead)):
            if newhead[i] =='NaN':
                pass
            else:
                sndhead[k] = newhead[i]
                k = k + 1
        #print sndhead
    fullhead = fullhead.tolist()
    return sndhead, fullhead

def LLD(geochem_data, element_list, imputation = False):
    '''
    This function is designed to find values that are below the detection
    limit of the analytical method. The function works by searching for the <
    symbol in the

    Parameters
    ----------
    geochem_data : dataframe
        Dataframe containg the geochemical data.
    element_List : TYPE
        DESCRIPTION.
    imputation : boolean, optional
        DESCRIPTION. The default is False.

    Returns
    -------
    The geochemistry .

    '''
    detection_limits = np.zeros([len(element_list)])
    for i in range (len(element_list)):
        less_than = False
        # find the values that have a < and turn that into a value
        if (geochem_data[element_list[i]].apply(str).\
                str.contains('<', na=False,regex=True).any()) == True:
            index = geochem_data[element_list[i]].\
                str.contains('<', na=False,regex=True)
            less_than = True
        else:
            index = geochem_data[element_list[i]].apply(str).\
                str.contains('-', na=False,regex=True)
            #except:
               # pass
        detection_index = geochem_data[element_list[i]].loc[index].unique()
        detection_index = list(detection_index)
        if len(detection_index) > 0:
            detection_index = detection_index[0]
            if less_than == True:
                new_detection_index = detection_index.replace('<','')
                new_detection_index = float(new_detection_index)
            else:
                new_detection_index = detection_index
            if new_detection_index <0:
                dl_value = new_detection_index*-1
            else:
                dl_value = new_detection_index
            #create a half detection limit value
            detection_limits[i] = dl_value
            half_detection = float(dl_value)/2
            # replace the less than values with half detection limits
            geochem_data[element_list[i]] = \
                geochem_data[element_list[i]].replace(detection_index,
                                                      half_detection)
    if imputation == True:
        print('Still to be implemented')
    return geochem_data, detection_limits

def repeats(geochem_data):
    '''
    Function used to find the location of repeats within a dataframe usign a
    key. The pair is presumed to be the location -1.

    Parameters
    ----------
    geochem_data : dataframe
        The dataframe containg the full geochemical dataset.

    Returns
    -------
    rep_location : list
        The location of the repeats as identified by the key.
    rep_pair : list
        The location of the corrisponding pair for each of the indentified
        repeats.

    '''
    # Find the location of repeats and their pair
    try:
        repeat = geochem_data[geochem_data[Id_Coloumn].str.contains(" R",
                                                                     na=False)]
    except KeyError:
        print ("Unable to find a column with the name: ", Id_Coloumn)
        sys.exit()
    rep_location = list(repeat.index)
    rep_pair = [x-1 for x in rep_location]
    return rep_location, rep_pair

def standard_stats(geochem_data, element_List, detection_limits = False):
    '''
    This Function is used to calculate the summary statistics for the analysed
    standards. The main statistics calculated are: mean, standard deviaiton,
    RSD (relative standard deviation). For analyses with more than one
    standard, the weighted average of the rsds is calculated using the
    mean concentration as as  weighting system.

    Parameters
    ----------
    geochem_data : Pandas dataframe
        Dataframe containg the geochemical data.
    element_List : list
        list containing the elements present within the dataframe.

    Returns
    -------
    None.

    '''
    stds_count = geochem_data[Id_Coloumn].value_counts()
    print(stds_count)
    Stds = stds_count[stds_count > STANDARD_CUTOFF]
    Stds_List = Stds.index.tolist()
    # Gernerate the headers for the Standard statistic
    for i in range (0, len(Stds_List)):
        if i == 0:
            stats_header = Stds_List[i] + '_Average'
        else:
            stats_header = np.append(stats_header,str(Stds_List[i]) +
                                     '_Average')
        stats_header = np.append(stats_header, [str(Stds_List[i]) + '_StDev',
                                                str(Stds_List[i]) +
                                                '_rsd(%)'])
    # Create a pandas dataframe to store the statitics
    # loop to add in a weighted average for more than one standard
    if len(Stds_List) >1:
        stats_header = np.append(stats_header, 'Weighted_Average')
        temp_stats = np.zeros([len(element_List),(len(Stds_List)*3)+1])
        average_locations = np.arange(0,len(Stds_List)*3, step = 3)
        rstdev_locations = [x+2 for x in average_locations]
    else:
        temp_stats = np.zeros([len(element_List),len(Stds_List)*3])
    std_stats = pd.DataFrame(data=temp_stats,
                             index=element_List,
                             columns= stats_header)
    pdf = matplotlib.backends.backend_pdf.PdfPages(Save_Location +
                                                   "\\Standards.pdf")
    try:
        #create a folder for the standrd images
        os.mkdir(Save_Location + "\\Standards")
    except FileExistsError:
        print ("Folder already exists")
    colour = ('tab:blue', 'tab:orange', 'tab:green', 'tab:red', 'tab:purple',
              'tab:cyan','tab:gray','tab:pink','tab:olive', 'dodgerblue')
    print (Stds_List)
    for j in range(0, len(element_List)):
        xlimit = 0
        y_upper = 0
        y_lower = 0
        for i in range(0, len(Stds_List)):
            element = geochem_data.loc[geochem_data[Id_Coloumn] ==
                                       Stds_List[i],element_List[j]]
            if min(element) < y_lower:
                y_lower = min(element)
            if max(element) > y_upper:
                y_upper = max(element)
            #option to turn on/off
            #element.index = np.arange(start=1, stop=(len(element)+1), step=1)
            element.columns = Stds_List[i]
            std_stats.loc[element_List[j],
                          stats_header[i+(2*i)]] = element.mean()
            '''
            need to look into if the mean is used and what impact it is having on the data

            '''
            std_stats.loc[element_List[j],
                          stats_header[(i+(2*i))+1]] = element.std()
            std_stats.loc[element_List[j],
                          stats_header[(i+(2*i))+2]] = (element.std()/
                                                        element.median())*100
            splot = sns.scatterplot(data=element, color = colour[i],
                                    label =  Stds_List[i])
            if max(element.index) > xlimit:
                xlimit = (max(element.index))
            plt.plot((0,99999),(element.median(),element.median()),
                     linewidth=0.5, color = colour[i])
            plt.plot((0,99999),((element.median())*1.1,(element.median())*1.1),
                     linewidth=0.5,linestyle='dashed', color = colour[i])
            plt.plot((0,99999),((element.median())*0.9,(element.median())*0.9),
                     linewidth=0.5, linestyle='dashed', color = colour[i])
        savename = ('{}\\Standards\\{}_Standards.png').format(Save_Location,
                                                              element_List[j])
        lgd = splot.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
        plt.xlim(0,xlimit)
        if detection_limits[j] > 0:
            plt.plot((0,99999),(detection_limits[j],detection_limits[j]),
                     linestyle='dashed', linewidth=1, color = 'k')
        if BATCHED == 1:
            for i in range(len(BATCHES)):
                plt.plot((BATCHES[i] + 0.5, BATCHES[i] + 0.5), (-99999,99999),
                         color = 'k', linestyle='dashed', linewidth = 0.5)
            plt.ylim(y_lower*1.3, y_upper*1.3)
        plt.title(element_List[j])
        plt.savefig(savename, bbox_extra_artists=(lgd,),
                    bbox_inches='tight', format = 'png',
                    orientation='portrait',dpi = 900)
        pdf.savefig()
        plt.close()
        plt.clf()
    pdf.close()
    if len(Stds_List) >1:
        for i in range(0, len(element_List)):
            equation_temp = 0
            sum_stds = std_stats.loc[element_List[i],
                                     stats_header[average_locations]].sum()
            for j in range (0,len(average_locations)):
                avg = std_stats.loc[element_List[i],
                                    stats_header[average_locations[j]]]
                rsd = std_stats.loc[element_List[i],
                                    stats_header[rstdev_locations[j]]]
                equation_temp = equation_temp + (rsd*avg)
            std_stats.loc[element_List[i],
                          'Weighted_Average'] = equation_temp/sum_stds
    standards_save = ('{}\\Standard_Stats.xlsx').format(Save_Location)
    try:
        std_stats.to_excel(standards_save)
    except xlsxwriter.exceptions.FileCreateError:
        print ('''Unable to save, the file must already be open. A new verison
               will try to be saved with _new. If this fails no file will be
               saved.''')
        standards_save = ('{}\\Standard_Stats_new.xlsx').format(Save_Location)
        std_stats.to_excel(standards_save)
    del pdf


def lab_assessment(geochem_data, element_list, detection_limits):
    '''
    We need to calculate Min, MAX, median and 90th percentile RSD (plus number
    ) for pairs for the following intervals:
    1-5* DL - group 1
    5-10* DL - group 2
    10-50* DL - group 3
    50-100* DL - group 4
    100-500* Dl - group 5
    500-1000* DL- group 6
    >1000* DL - group 7


    Parameters
    ----------
    geochem_data : TYPE
        DESCRIPTION.
    elements_list : TYPE
        DESCRIPTION.

    Returns
    -------
    None.

    '''
    try:
        detection_limits  = pd.read_excel(detection_limits, header = 0)
    except FileNotFoundError:
        print ('Could not find a file with that name')
        sys.exit()
    #count if to determine groups
    repeat_loation, repeat_Pair = repeats(geochem_data)
    dl = detection_limits
    detection_limits = detection_limits[element_list]
    dl = list(detection_limits)
    pdf = matplotlib.backends.backend_pdf.PdfPages(Save_Location +
                                               "\\Laboratory_Assessment.pdf")
    try:
        os.mkdir(Save_Location + "\\Laboratory_Assessment")
    except FileExistsError:
        print ("Folder already exists")
    for i in range(0, len(element_list)):
        dl = detection_limits[element_list[i]]
        dl = list(dl)
        x = geochem_data.iloc[repeat_loation,list(
            geochem_data.columns).index(element_list[i])]
        y = geochem_data.iloc[repeat_Pair,list(
            geochem_data.columns).index(element_list[i])]
        means = np.zeros([len(x)])
        rpt = list(x)
        rpt_pair = list(y)
        for j in range(len(x)):
            means[j] = np.mean((rpt[j], rpt_pair[j]))
        #find the number of samples in each group
        group_count = np.zeros([7])
        group_ranges = [1,5,10,50,100,500,1000,9999999999]
        for j in range(7):
            group_count[j] = sum(1 for x in means if x> group_ranges[j]*dl[0]
                                 and x< group_ranges[j+1]*dl[0])
        temp_summary_stats = np.zeros([4,7])
        stats_header = ('1-5* DL', '5-10* DL', '10-50* DL', '50-100* DL',
                        '100-500* Dl', '500-1000* DL', '>1000* DL')
        stats_index = ('Minimum','Maximum','Median','90th percentile')
        summary_stats = pd.DataFrame(data=temp_summary_stats,
                                     index=stats_index,columns= stats_header)
        for j in range(7):
            if group_count[j] > 1:
                repeat  = np.array(rpt)
                repeat_pair = np.array(rpt_pair)
                repeat_index = [x for x, y in enumerate(means) if
                                y> group_ranges[j]*dl[0]and
                                y< group_ranges[j+1]*dl[0]]
                if DEBUG == 1:
                    print (group_count[j], 'number in group')
                    print (len(repeat_index), 'number found')
                    print (group_ranges[j]*dl[0], group_ranges[j+1]*dl[0])
                    print (dl, 'detection limit')
                    print (means[repeat_index], 'found values')
                    print (repeat[repeat_index], 'repeats')
                    print (repeat_pair[repeat_index], 'repeat pairs')
                    print ('---------------------')
                #summary_stats[stats_header[j]] = duplicate_statistics(
                    #repeat[repeat_index], repeat_pair[repeat_index])
                #minimum, maximum, median, percentile =  duplicate_statistics(
                    #repeat[repeat_index], repeat_pair[repeat_index])
            else:
                summary_stats[stats_header[j]] = ['NaN','NaN','NaN','NaN']
        stats_save = ('{}\\Laboratory_Assessment_Stats.xlsx').format(
            Save_Location)
        if i == 0:
            with pd.ExcelWriter(stats_save, engine='xlsxwriter') as writer:
                summary_stats.to_excel(writer, element_list[i])
                writer.save()
        book = load_workbook(stats_save)
        with pd.ExcelWriter(stats_save, engine='openpyxl') as writer:
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            summary_stats.to_excel(writer, element_list[i])
            writer.save()
        plt.scatter(x, y, s= 3)
        plt.title(element_list[i])
        colour = ['r','cornflowerblue','forestgreen','orange',
                  'teal','darkmagenta','k']
        for count in range(7):
            detection_line = group_ranges[count]*dl[0]
            plt.plot([detection_line,detection_line],[0,999999],
                     linestyle='dashed', color = colour[count])
            plt.plot([0,999999],[detection_line,detection_line],
                     linestyle='dashed', color = colour[count])
        x_min = min(x)-(min(x)*0.2)
        y_min = min(y)-(min(y)*0.2)
        if x_min <0:
            x_min = 0
        if y_min <0:
            y_min = 0
        plt.xlim(x_min,max(x)+(max(x)*0.2))
        plt.ylim(y_min,max(y)+(max(y)*0.2))
        plt.yscale('log')
        plt.xscale('log')
        savename = ('{}\\Laboratory_Assessment\\{}_Laboratory_Assessment.png'
                    ).format(Save_Location,element_list[i])
        plt.savefig(savename, format = 'png', orientation='portrait',
                    dpi = 900)
        pdf.savefig()
        plt.close()
        plt.clf()
    writer.close()
    pdf.close()
                #for count in range(group_count[j]):

                #use repeat and repeat pair and duplicate_statistics function

def Duplicates(geochem_data, element_list, key_word, SheetName):
    repeat = geochem_data[geochem_data[Id_Coloumn].str.contains(key_word,na=False)]
    rep_location = list(repeat.index)
    repeat_list = list(repeat[Id_Coloumn])
    repeat_list = [item.replace(key_word, "") for item in repeat_list]
    duplicates = pd.DataFrame(np.zeros((len(rep_location)*2, len(list(geochem_data)))),columns = list(geochem_data))
    #duplicates.loc[0] = geochem_data.loc[0]
    for i in range(0,len(rep_location)):
        duplicates.loc[i*2] = geochem_data.loc[rep_location[i]]
        try:
            repeat_location = geochem_data[geochem_data[Id_Coloumn].str.contains(repeat_list[i],na=False)]
            try:
                repeat_location = geochem_data[(geochem_data[Id_Coloumn]) == int(repeat_list[i])]
            except:
                repeat_location = geochem_data[(geochem_data[Id_Coloumn]) == str(repeat_list[i])]
            repeat_location = (list(repeat_location.index))
            duplicates.loc[(i*2)+1] = geochem_data.loc[repeat_location[0]]
        except:
            pass
    duplicate_save = ('{}\\{}_Duplicates.xlsx').format(Save_Location, BATCH)
    file_exists = os.path.isfile(duplicate_save)
    if file_exists == False:
        duplicates.to_excel(duplicate_save, index = False,sheet_name = SheetName)
        book = load_workbook(duplicate_save)
        writer = pd.ExcelWriter(duplicate_save, engine = 'openpyxl')
        writer.book = book
    else:
        book = load_workbook(duplicate_save)
        writer = pd.ExcelWriter(duplicate_save, engine = 'openpyxl')
        writer.book = book
        duplicates.to_excel(writer, index = False,sheet_name = SheetName)
    stats = Duplicate_Statistics(duplicates)
    stats.to_excel(writer, index = False,sheet_name = SheetName+"_Statistics")
    writer.save()
    writer.close()

def Duplicate_Statistics(duplicates):
    header = list(duplicates)
    element_list, duplicates.columns = parse(header)
    number_duplicates = int(len(duplicates)/2)
    stats = np.zeros([number_duplicates, 4])
    duplicate_stats = pd.DataFrame(np.zeros([number_duplicates*5,len(element_list)+1]))
    header_info = np.append("",element_list)
    duplicate_stats.columns = header_info
    for j in range(len(element_list)):
        for i in range(number_duplicates):
            repeat = duplicates[element_list[j]].iloc[i*2]
            pair = duplicates[element_list[j]].iloc[(i*2)+1]
            stats[i,0] = st.mean([repeat, pair])
            stats[i,1] = st.stdev([repeat, pair])
            dup_id = duplicates.columns.get_loc(Id_Coloumn)
            duplicate_stats.iloc[(i*5),0] = duplicates.iloc[i*2,dup_id]
            duplicate_stats.iloc[(i*5)+1,0] =  "Mean"
            duplicate_stats.iloc[(i*5)+2,0] = "Standard Deviation"
            duplicate_stats.iloc[(i*5)+3,0] = "RSD"
            duplicate_stats.iloc[(i*5)+4,0] = "RPD"
            duplicate_stats.iloc[(i*5),j+1] = ""
            duplicate_stats.iloc[(i*5)+1,j+1] =  st.mean([repeat, pair])
            duplicate_stats.iloc[(i*5)+2,j+1] = st.stdev([repeat, pair])
            duplicate_stats.iloc[(i*5)+3,j+1] = stats[i,1]/stats[i,0]
            duplicate_stats.iloc[(i*5)+4,j+1] = abs(((repeat-pair)/((repeat+pair)/2))*100) #Relative Percent Difference
    return duplicate_stats

def main():
    try:
        geochem_data  = pd.read_excel(FILE_NAME, header = 0)
    except FileNotFoundError:
        print ('Could not find a file with that name')
        sys.exit()
    #
    header = list(geochem_data)
    element_list, geochem_data.columns = parse(header)
    geochem_data_unstripped = geochem_data.copy(deep = True)
    geochem_data[Id_Coloumn] = geochem_data[Id_Coloumn].replace(' Rpt', '',
                                                               regex=True)
    # creates the plots and runs standard statistics
    #print (geochem_data.loc[(250)])
    geochem_data, detection_limits = LLD(geochem_data,element_list)
    print(detection_limits)
    print (element_list)
    standard_stats(geochem_data,element_list, detection_limits)
    Duplicates(geochem_data, element_list, " DUP", "Lab_Duplicates")
    Duplicates(geochem_data_unstripped, element_list, " Rpt", "Analytical")
    #creates the plots for duplicate pairs
    #duplicates(geochem_data_unstripped, element_list)
    #lab_assessment(geochem_data_unstripped, element_list, detection_limits)

if __name__ == "__main__":
    main()